import json
import os
from enum import Enum, auto
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import numpy as np
from typing import Any

class AircraftType(Enum):
    JET = auto()
    PROP = auto()
    MIXED = auto()

class FlapType(Enum):
    PLAIN_SPLIT = auto()
    SLOT = auto()
    FOWLER = auto()
    DOUBLE_SLOT = auto()
    TRIPLE_SLOT = auto()
class StressOutput(Enum):
    DEFLECTION = auto()
    TWIST = auto()
    BENDING_STRESS = auto()
    BENDING_STRESS_BOTTOM = auto()
    SHEAR_STRESS = auto()
    TORSION = auto()
    RESULTANT_VERTICAL = auto()
    RESULTANT_HORIZONTAL = auto()
    INTERNAL_SHEAR_VERTICAL = auto()
    INTERNAL_SHEAR_HORIZONTAL = auto()
    INTERNAL_MOMENT_X = auto()
    INTERNAL_MOMENT_Y = auto()
    INTERNAL_TORQUE = auto()
    SHEAR_STRESS_TOP = auto()
    SHEAR_STRESS_BOTTOM = auto()
    WING_BENDING_STRESS = auto()
    RESULTANT_TORQUE = auto()

class EvaluateType(Enum):
    VERTICAL = auto()
    WING = auto()
    HORIZONTAL = auto()
class WingType(Enum):
    HIGH = auto()
    LOW = auto()
    
class LoadCase(Enum):
    OEW = auto()
    OEW_PAYLOAD = auto()
    OEW_PAYLOAD_FUEL = auto()
    OEW_FUEL = auto()

class EmpType(Enum):
    CRUCIFORM = auto()
    T_TAIL = auto()
    CONVENTIONAL = auto()
    H_TAIL = auto()
    NONE = auto()

class MissionType(Enum):
    DESIGN = auto()
    FERRY = auto()
    ALTITUDE = auto()

class AircraftType(Enum):
    JET = auto()
    PROP = auto()
    MIXED = auto()

class MissionType(Enum):
    DESIGN = auto()
    FERRY = auto()
    ALTITUDE = auto()

class Materials(Enum):
    Al7075 = auto()
    Al6061 = auto()
    Al6063 = auto()
    Al2024 = auto()
    Al5052 = auto()
    Ti10V2Fe3Al = auto()
    Ti6Al4V = auto()
    INCONEL_718 = auto()
    WASPALOY = auto()
    Rene_41 = auto()
    HASTELLOY_X = auto()

class EnumEncoder(json.JSONEncoder):
    """
    Makes sure that enums are saved as strings correctly in the json file
    """
    def default(self, o):
        if isinstance(o, Enum):
            return o.name
        return json.JSONEncoder.default(self, o)


class Data:
    def __init__(self, design_file, file_type='design'):
        if file_type == 'design':
            self.data = self.load_design(design_file)
        elif file_type == 'airfoil_geometry':
            self.data = self.load_dat_file(design_file)
        elif file_type == 'aerodynamics':
            self.data = self.load_aerodynamic_distribution(design_file)

    def load_design(self, design_file) -> dict[str, Any]:
        file_path = os.path.join("Data", design_file)
        try:
            with open(file_path, 'r') as file:
                return json.load(file)
        except FileNotFoundError:
            print(f"Error: {file_path} not found.")
        except json.JSONDecodeError:
            print(f"Error: Failed to decode JSON from {file_path}.")
        return {}

    def save_design(self, design_file):
        file_path = os.path.join("Data", design_file)
        try:
            with open(file_path, 'w') as file:
                json.dump(self.data, file, cls=EnumEncoder, indent=4)
        except IOError as e:
            print(f"Error: Failed to write to {file_path}. {e}")

    def load_dat_file(self, dat_file) -> dict[str, list[float]]:
        """
        Load a .dat file as a dictionary with 'x' and 'y' keys.
        Assumes each line contains two numeric values separated by space or comma.
        """
        file_path = os.path.join("Data", dat_file)
        x_vals, y_vals = [], []
        try:
            with open(file_path, 'r') as file:
                for line in file:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.replace(',', ' ').split()
                    if len(parts) != 2:
                        print(f"Warning: Skipping malformed line: {line}")
                        continue
                    try:
                        x, y = float(parts[0]), float(parts[1])
                        x_vals.append(x)
                        y_vals.append(y)
                    except ValueError:
                        print(f"Warning: Non-numeric data encountered in line: {line}")
            if any(i for i in y_vals if i < 0):
                y_vals = [i + max(y_vals) for i in y_vals] 
            return {"x": x_vals, "y": y_vals}
        except FileNotFoundError:
            print(f"Error: {file_path} not found.")
        except IOError as e:
            print(f"Error: Failed to read from {file_path}. {e}")
        return {}

    def load_aerodynamic_distribution(self, filename: str):
        
        folder_path = os.path.join(os.path.dirname(__file__), "Data")
        file_path = os.path.join(folder_path, filename)

        yspan_values = []
        cl_values = []
        chord_values = []
        induced_cd_values = []
        Cm_values = []
        header_found = False

        with open(file_path, 'r') as f:
            for line in f:
                stripped = line.strip()
                # Look for table header by checking for both 'y-span' and 'Cl'
                if not header_found and "y-span" in stripped and "Cl" in stripped and "Chord" in stripped:
                    header_found = True
                    continue  # skip the header line
                if header_found:
                    # Stop reading if we hit an empty line
                    if not stripped:
                        break
                    # Split by whitespace. This assumes the columns are space-separated.
                    parts = stripped.split()
                    if len(parts) < 4:
                        continue
                    try:
                        y_span = float(parts[0])
                        cl = float(parts[3])
                        chord = float(parts[1])
                        Icd =float(parts[5])
                        Cm = float(parts[7])
                        yspan_values.append(y_span)
                        cl_values.append(cl)
                        chord_values.append(chord)
                        induced_cd_values.append(Icd)
                        Cm_values.append(Cm)
                    except ValueError:
                        # In case conversion fails, skip the row
                        continue
            
            forces_dict = {
                "yspan": yspan_values,
                "cl": cl_values,
                "chord": chord_values,
                "induced_cd": induced_cd_values,
                "cm": Cm_values
            }

            return forces_dict


def generate_df():
    all_rows = []
    for i in range(1, 5):
        file_path = os.path.join(os.path.dirname(__file__), "Data", f"design{i}.json")
        if not os.path.exists(file_path):
            print(f"File {file_path} does not exist. Skipping.")
            continue

        aircraft_data = Data(file_path)
        base_data = aircraft_data.data.copy()

        mission_keys = {"design", "ferry", "altitude"}
        static_data = {k: v for k, v in base_data.items() if k not in mission_keys and not isinstance(v, dict)}

        for mission in MissionType:
            mission_name = mission.name.lower()
            if mission_name not in base_data:
                continue

            mission_data = base_data[mission_name]
            row = {**static_data, **mission_data}
            row["mission_type"] = mission_name
            row["design_id"] = base_data.get("design_id", i)
            all_rows.append(row)

    df = pd.DataFrame(all_rows)
    return df

def kg2lbs(kg):
    """
    Convert kg to lbs
    """
    return kg * 2.2046226218487757

def lbs2kg(lbs):
    """
    Convert lbs to kg
    """
    return lbs / 2.2046226218487757

def N2lbf(N):
    """
    Convert N to lbf
    """
    return N * 0.224809

def lbf2N(lbf):
    """
    Convert lbf to N
    """
    return lbf / 0.224809

def ft2m(ft):
    """
    Convert ft to m
    """
    return ft * 0.3048

def m2ft(m):
    """
    Convert m to ft
    """
    return m / 0.3048

def msq2ftsq(m2):
    """
    Convert m^2 to ft^2
    """
    return m2 * 10.7639

def ftsq2msq(ft2):
    """
    Convert ft^2 to m^2
    """
    return ft2 / 10.7639

def rad2deg(rad):
    """
    Convert radians to degrees
    """
    return rad * (180 / np.pi)

def deg2rad(deg):
    """
    Convert degrees to radians
    """
    return deg * (np.pi / 180)

def kgperm32lbsperft3(kgm3):
    """
    Convert kg/m^3 to lbs/ft^3
    """
    return kgm3 * 0.062428

def lbsperft32kgperm3(lbsft3):
    """
    Convert lbs/ft^3 to kg/m^3
    """
    return lbsft3 / 0.062428

def Pa2lbfpftsq(Pa):
    """
    Convert Pa to lb/ft^2
    """
    return Pa * 0.0208854

def lbfpftsq2Pa(lbft2):
    """
    Convert lb/ft^2 to Pa
    """
    return lbft2 / 0.0208854

def kgpJ2lbsphrphp(kgpJ):
    """
    Convert kg/(J/s) to lbs/(hp/hr)
    """
    # TODO: Check if this conversion is correct i dont think so - Owen
    return kgpJ * 0.02388458966275

def deg2rad(deg):
    """
    Convert degrees to radians
    """
    return deg * (np.pi / 180)

def rad2deg(rad):
    """
    Convert radians to degrees
    """
    return rad * (180 / np.pi)

def L2gal(L):
    """
    Convert Liters to gallons
    """
    return L * 0.264172

def gal2L(gal):
    """
    Convert gallons to Liters
    """
    return gal / 0.264172

def kgmsq2lbsftsq(kgm2):
    """
    Convert kg/m^2 to lbs/ft^2
    """
    return kgm2 * 0.204816

def lbsftsq2kgmsq(lbsft2):
    """
    Convert lbs/ft^2 to kg/m^2
    """
    return lbsft2 / 0.204816

def W2hp(W):
    """
    Convert Watts to horsepower
    """
    return W * 0.00134102

def hp2W(hp):
    """
    Convert horsepower to Watts
    """
    return hp / 0.00134102

def apply_number_format(cell, value):
    if isinstance(value, (int, float)):
        if abs(value) >= 0.001:
            cell.number_format = '#,##0.000'  # Locale-dependent: assumes Excel settings use comma for decimals
        elif value == 0:
            cell.number_format = '0'
        else:
            cell.number_format = '0.000E+00'
        cell.value = value
    else:
        cell.value = value

def auto_adjust_column_widths(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value)
                if value:
                    max_length = max(max_length, len(value))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

def design_json_to_excel(json_file: str, excel_file: str) -> None:
    data = Data(json_file).data

    variable_units = {
        "design_range": "[m]", "design_payload": "[kg]", "ferry_range": "[m]", "reserve_range": "[m]",
        "ferry_payload": "[kg]", "altitude_range_WIG": "[m]", "altitude_range_WOG": "[m]", "altitude_payload": "[kg]",
        "cruise_speed": "[m/s]", "jet_consumption": "[kg/(N·s)]", "prop_consumption": "[kg/J]",
        "reserve_fuel": "[N]", "take_off_power": "[W]", "take_off_thrust": "[N]", "cruise_altitude": "[m]",
        "MTOM": "[kg]", "MTOW": "[N]", "OEW": "[N]", "ZFW": "[N]", "EW": "[N]", "max_fuel": "[N]",
        "mission_fuel": "[N]", "S": "[m²]", "b": "[m]", "MAC": "[m]", "WP": "[N/W]",
        "WS": "[N/m²]", "fuel_economy": "[L/ton/km]", "P": "[W]", "T": "[N]", "stall_speed_clean": "[m/s]",
        "stall_speed_takeoff": "[m/s]", "stall_speed_landing": "[m/s]", "high_altitude": "[m]", "L": "[m]", "r": "[m]",
        "hull_surface": "[m²]", "gravitational_acceleration": "[m/s²]", "kinematic_viscosity": "[m²/s]", "viscosity_air": "[m²/s]", "rho_water": "[kg/m³]", 
        "l_fuselage": "[m]", "l_cargo_straight": "[m]", "r_fuselage": "[m]", "d_fuselage": "[m]", "l_tailcone": "[m]", "l_nose": "[m]",
        "V_lof": "[m/s]", "cargo_width": "[m]", "cargo_height": "[m]", "cargo_length": "[m]", "cargo_density": "[kg/m³]",
        "sweep_c_4": "[deg]", "dihedral": "[deg]", "sweep_x_c": "[deg]", "sweep_TE": "[deg]", "chord_root": "[m]", "chord_tip": "[m]",
        "y_MAC": "[m]", "X_LEMAC": "[m]", "X_LE": "[m]", "total_fuel": "[N]", "rho_air": "[kg/m³]", "S_h": "[m²]", 
        "S_v": "[m²]", "l_h": "[m]", "l_v": "[m]", "most_aft_cg": "[m]", "most_forward_cg": "[m]", "sweep": "[deg]",
        "r_float": "[m]", "upsweep": "[deg]", "LE_pos": "[m]", "mission_fuel_L": "[L]", "total_fuel_L": "[L]", "max_fuel_L": "[L]", "reserve_fuel_L": "[L]"
    }

    requirements = data.get('requirements', {})
    inputs = data.get('inputs', {})
    general_outputs = data.get('outputs', {}).get('general', {})
    design_outputs = data.get('outputs', {}).get('design', {})
    max_outputs = data.get('outputs', {}).get('max', {})
    wing_design = data.get('outputs', {}).get('wing_design', {})
    vertical_tail = data.get('outputs', {}).get('empennage_design', {}).get('vertical_tail',{})
    horizontal_tail = data.get('outputs', {}).get('empennage_design', {}).get('horizontal_tail', {})
    cg_range = data.get('outputs', {}).get('cg_range', {})

    del data['outputs']
    del data['requirements']
    del data['inputs']

    try:
        excelsheet = openpyxl.load_workbook(excel_file)
        if f"Design {data['design_id']} Data" in excelsheet.sheetnames:
            del excelsheet[f"Design {data['design_id']} Data"]
        sheet = excelsheet.create_sheet(title=f"Design {data['design_id']} Data")
    except FileNotFoundError:
        excelsheet = openpyxl.Workbook()
        sheet = excelsheet.active
        sheet.title = f"Design {data['design_id']} Data"

    bold_font = Font(bold=True)

    def write_block(start_row, col1, col2, title, block_data):
        sheet.cell(row=start_row, column=col1, value=f"Design {data['design_id']}").font = bold_font
        sheet.cell(row=start_row, column=col2, value=title).font = bold_font
        row = start_row + 1
        for key, value in block_data.items(): 
            if title == "Inputs" and key in {"h_position", "v_position", "sweep_h", "sweep_v", "aspect_h", "aspect_v", "sweep_hc4", "taper_h", "taper_v"}:
                continue
            label = f"{key} {variable_units[key]}" if key in variable_units else key
            sheet.cell(row=row, column=col1, value=label)
            apply_number_format(sheet.cell(row=row, column=col2), value)
            row += 1
        return row

    row_req_end = write_block(1, 1, 2, "Requirements", requirements)
    row_gen_states_end = write_block(row_req_end + 1, 1, 2, "General States", data)
    row_inputs_end = write_block(1, 4, 5, "Inputs", inputs)
    row_general_out_end = write_block(1, 7, 8, "General Outputs", general_outputs)

    sheet.cell(row=1, column=10, value=f"Design {data['design_id']}").font = bold_font
    sheet.cell(row=1, column=11, value="Design Mission").font = bold_font

    excluded_keys = {"S", "b", "MAC", "MTOM", "fuel_economy", "LD", "Mff", "max_fuel", "total_fuel", "mission_fuel", "reserve_fuel"}
    row = 2
    for key, value in design_outputs.items():
        if key in excluded_keys:
            continue
        label = f"{key} {variable_units[key]}" if key in variable_units else key
        sheet.cell(row=row, column=10, value=label)
        apply_number_format(sheet.cell(row=row, column=11), value)
        row += 1

    for key, value in max_outputs.items():
        label = f"{key} {variable_units[key]}" if key in variable_units else key
        sheet.cell(row=row, column=10, value=label)
        apply_number_format(sheet.cell(row=row, column=11), value)
        row += 1

    write_block(row_inputs_end + 1, 4, 5, "Wing Design", wing_design)
    horizontal_end = write_block(row_general_out_end + 1, 7, 8, "Horizontal Tail", horizontal_tail)
    vertical_end = write_block(horizontal_end + 1, 7, 8, "Vertical Tail", vertical_tail)

    write_block(vertical_end + 1, 7, 8, "CG Range", cg_range)

    auto_adjust_column_widths(sheet)

    excelsheet.save(excel_file)

class ISA:
    """
    A class representing the International Standard Atmosphere (ISA) model.

    This class calculates atmospheric properties (pressure, temperature, 
    density, and speed of sound) based on altitude using the standard ISA model.

    Attributes:
        altitude (float): Altitude in meters.
        pressure (float): Air pressure at the given altitude in Pascals.
        temperature (float): Air temperature at the given altitude in Kelvin.
        rho (float): Air density at the given altitude in kg/m^3.
        speed_of_sound (float): Speed of sound at the given altitude in m/s.
    
    Methods:
        Mach(IAS): Computes the Mach number for a given Indicated Airspeed (IAS).
        TAS(IAS): Computes the True Airspeed (TAS) for a given IAS.
    """
    def __init__(self, altitude: float=0) -> None:
        """
        Initialize the ISA model for a given altitude.

        Args:
            altitude (float): Altitude in meters.
        """
        self.lamb = -0.0065  # Temperature lapse rate in K/m
        self.Temp0 = 288.15  # Sea level standard temperature in K
        self.p0 = 101325  # Sea level standard pressure in Pa
        self.rho0 = 1.225  # Sea level standard density in kg/m^3
        self.gamma = 1.4  # Ratio of specific heats for air
        self.R = 287.05  # Specific gas constant for dry air in J/(kg·K)
        self.g = 9.80665  # Acceleration due to gravity in m/s^2
        
        self.altitude = altitude


    @property
    def altitude(self) -> float:
        return self._altitude

    @altitude.setter
    def altitude(self, value: float) -> None:
        if value < 0:
            raise ValueError("Altitude cannot be negative")
        self._altitude = value
        self._pressure = self._compute_pressure()
        self._temperature = self._compute_temperature()
        self._rho = self._compute_rho()
        self._speed_of_sound = self._compute_speed_of_sound()

    def _compute_pressure(self) -> float:
        power = -self.g / (self.lamb * self.R)
        return self.p0 * (1 + self.lamb * self.altitude / self.Temp0) ** power

    def _compute_temperature(self) -> float:
        return self.Temp0 + self.lamb * self.altitude
    
    def _compute_rho(self) -> float:
        return self._pressure / (self.R * self._temperature)
    
    def _compute_speed_of_sound(self) -> float:
        return np.sqrt(self.gamma * self.R * self._temperature)
    
    @property
    def pressure(self) -> float:
        return self._pressure

    @property
    def temperature(self) -> float:
        return self._temperature
    
    @property
    def rho(self) -> float:
        return self._rho
    
    @property
    def speed_of_sound(self) -> float:
        return self._speed_of_sound
    
    def Mach(self, IAS: float) -> float:
        """
        Calculate the Mach number given Indicated Airspeed (IAS).
        ---------------------------------------------------------
        Parameters:
            IAS : float
                Indicated airspeed in m/s.

        Returns:
            float: Mach number.
        """
        curly_braces = (1 + ((self.gamma-1)/(2*self.gamma)) * self.rho0/self.p0 * IAS ** 2) ** (self.gamma/(self.gamma-1)) - 1 
        square_braces = (1 + self.p0/self.pressure * curly_braces) ** ((self.gamma-1)/self.gamma) - 1
        M = np.sqrt(2/(self.gamma-1) * square_braces)
        return M

    def TAS(self, IAS: float) -> float:
        """
        Calculate the true airspeed (TAS) given Indicated Airspeed (IAS).
        ------------------------------------------------------------------
        Parameters:
            IAS (float): Indicated airspeed in m/s.

        Returns:
            float: True airspeed in m/s.
        """
        M = self.Mach(IAS)
        return M * self.speed_of_sound
    
    def TAT_to_SAT(self, IAS: float, measured_TAT: float) -> float:
        """
        Calculate the static air temperature (SAT) given Indicated Airspeed (IAS).
        ---------------------------------------------------------------------------
        Parameters:
            IAS (float): Indicated airspeed in m/s.
            measured_TAT (float): Total air temperature in Kelvin (measured).

        Returns:
            float: Static air temperature in Kelvin.
        """
        M = self.Mach(IAS)
        return measured_TAT / (1 + (self.gamma-1)/2 * M ** 2)

    def EAS(self, TAS: float) -> float:
        """
        Calculate the equivalent airspeed (EAS) given True Airspeed (TAS).
        -------------------------------------------------------------------
        Parameters:
            TAS (float): True airspeed in m/s.
            rho (float): Air density at the given altitude in kg/m^3.

        Returns:
            float: Equivalent airspeed in m/s.
        """
        return TAS * np.sqrt(self.rho / self.rho0)

# Example usage
if __name__ == "__main__":
    isa = ISA(altitude=ft2m(10000))
    print("Pressure at 10,000 ft:", isa.pressure, "Pa")
    # aircraft = Data("design1.json")
    # print(aircraft.data)
    # aircraft.data["aircraft_type"] = AircraftType.JET.name
    # aircraft.save_design("design1.json")
    # print(aircraft.data)

    # print("Aircraft Type:", AircraftType[aircraft.data["aircraft_type"]])