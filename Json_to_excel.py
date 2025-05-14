from utils import Data
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def apply_number_format(cell, value):
    if isinstance(value, (int, float)):
        if abs(value) >= 0.001:
            cell.number_format = '#,##0.000'  # Locale-dependent: assumes Excel settings use comma for decimals
        elif value == 0:
            cell.number_format = '0'
        else:
            cell.number_format = '0.000E+00'
        cell.value = value
    else:
        cell.value = value

def auto_adjust_column_widths(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value)
                if value:
                    max_length = max(max_length, len(value))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

def design_json_to_excel(json_file: str, excel_file: str) -> None:
    data = Data(json_file).data

    variable_units = {
        "design_range": "[m]", "design_payload": "[kg]", "ferry_range": "[m]", "reserve_range": "[m]",
        "ferry_payload": "[kg]", "altitude_range_WIG": "[m]", "altitude_range_WOG": "[m]", "altitude_payload": "[kg]",
        "cruise_speed": "[m/s]", "jet_consumption": "[kg/(N·s)]", "prop_consumption": "[kg/J]",
        "reserve_fuel": "[N]", "take_off_power": "[W]", "take_off_thrust": "[N]", "cruise_altitude": "[m]",
        "MTOM": "[kg]", "MTOW": "[N]", "OEW": "[N]", "ZFW": "[N]", "EW": "[N]", "max_fuel": "[N]",
        "mission_fuel": "[N]", "S": "[m²]", "b": "[m]", "MAC": "[m]", "WP": "[N/W]",
        "WS": "[N/m²]", "fuel_economy": "[L/ton/km]", "P": "[W]", "T": "[N]", "stall_speed_clean": "[m/s]",
        "stall_speed_takeoff": "[m/s]", "stall_speed_landing": "[m/s]", "high_altitude": "[m]", "L": "[m]", "r": "[m]",
        "hull_surface": "[m²]", "gravitational_acceleration": "[m/s²]", "kinematic_viscosity": "[m²/s]", "viscosity_air": "[m²/s]", "rho_water": "[kg/m³]", 
        "l_fuselage": "[m]", "l_cargo_straight": "[m]", "r_fuselage": "[m]", "d_fuselage": "[m]", "l_tailcone": "[m]", "l_nose": "[m]",
        "V_lof": "[m/s]", "cargo_width": "[m]", "cargo_height": "[m]", "cargo_length": "[m]", "cargo_density": "[kg/m³]",
        "sweep_c_4": "[deg]", "dihedral": "[deg]", "sweep_x_c": "[deg]", "sweep_TE": "[deg]", "chord_root": "[m]", "chord_tip": "[m]",
        "y_MAC": "[m]", "X_LEMAC": "[m]", "X_LE": "[m]", "total_fuel": "[N]", "rho_air": "[kg/m³]", "S_h": "[m²]", 
        "S_v": "[m²]", "l_h": "[m]", "l_v": "[m]", "most_aft_cg": "[m]", "most_forward_cg": "[m]", "sweep": "[deg]",
        "r_float": "[m]", "upsweep": "[deg]"
    }

    requirements = data.get('requirements', {})
    inputs = data.get('inputs', {})
    general_outputs = data.get('outputs', {}).get('general', {})
    design_outputs = data.get('outputs', {}).get('design', {})
    max_outputs = data.get('outputs', {}).get('max', {})
    wing_design = data.get('outputs', {}).get('wing_design', {})
    vertical_tail = data.get('outputs', {}).get('empennage_design', {}).get('vertical_tail',{})
    horizontal_tail = data.get('outputs', {}).get('empennage_design', {}).get('horizontal_tail', {})
    cg_range = data.get('outputs', {}).get('cg_range', {})

    del data['outputs']
    del data['requirements']
    del data['inputs']

    try:
        excelsheet = openpyxl.load_workbook(excel_file)
        if f"Design {data['design_id']} Data" in excelsheet.sheetnames:
            del excelsheet[f"Design {data['design_id']} Data"]
        sheet = excelsheet.create_sheet(title=f"Design {data['design_id']} Data")
    except FileNotFoundError:
        excelsheet = openpyxl.Workbook()
        sheet = excelsheet.active
        sheet.title = f"Design {data['design_id']} Data"

    bold_font = Font(bold=True)

    def write_block(start_row, col1, col2, title, block_data):
        sheet.cell(row=start_row, column=col1, value=f"Design {data['design_id']}").font = bold_font
        sheet.cell(row=start_row, column=col2, value=title).font = bold_font
        row = start_row + 1
        for key, value in block_data.items(): 
            if title == "Inputs" and key in {"h_position", "v_position", "sweep_h", "sweep_v", "aspect_h", "aspect_v", "sweep_hc4", "taper_h", "taper_v"}:
                continue
            label = f"{key} {variable_units[key]}" if key in variable_units else key
            sheet.cell(row=row, column=col1, value=label)
            apply_number_format(sheet.cell(row=row, column=col2), value)
            row += 1
        return row

    row_req_end = write_block(1, 1, 2, "Requirements", requirements)
    row_gen_states_end = write_block(row_req_end + 1, 1, 2, "General States", data)
    row_inputs_end = write_block(1, 4, 5, "Inputs", inputs)
    row_general_out_end = write_block(1, 7, 8, "General Outputs", general_outputs)

    sheet.cell(row=1, column=10, value=f"Design {data['design_id']}").font = bold_font
    sheet.cell(row=1, column=11, value="Design Mission").font = bold_font

    excluded_keys = {"S", "b", "MAC", "MTOM", "fuel_economy"}
    row = 2
    for key, value in design_outputs.items():
        if key in excluded_keys:
            continue
        label = f"{key} {variable_units[key]}" if key in variable_units else key
        sheet.cell(row=row, column=10, value=label)
        apply_number_format(sheet.cell(row=row, column=11), value)
        row += 1

    for key, value in max_outputs.items():
        label = f"{key} {variable_units[key]}" if key in variable_units else key
        sheet.cell(row=row, column=10, value=label)
        apply_number_format(sheet.cell(row=row, column=11), value)
        row += 1

    write_block(row_inputs_end + 1, 4, 5, "Wing Design", wing_design)
    horizontal_end = write_block(row_general_out_end + 1, 7, 8, "Horizontal Tail", horizontal_tail)
    vertical_end = write_block(horizontal_end + 1, 7, 8, "Vertical Tail", vertical_tail)

    write_block(vertical_end + 1, 7, 8, "CG Range", cg_range)

    auto_adjust_column_widths(sheet)

    excelsheet.save(excel_file)

# Example call
if __name__ == "__main__":
    json_file = "design1.json"
    excel_file = "Concept Data.xlsx"
    design_json_to_excel(json_file, excel_file)
